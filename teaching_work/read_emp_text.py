import openpyxl
file = open("emp_data.txt", "r")
#content = file.readlines()
emp_list = list()
key = ['name', 'age', 'cName', 'salary', 'designation']
for line in file.readlines():
    dlist =  line.strip().split()
    d = dict()
    for i in range(len(dlist)):
        d[key[i]] = dlist[i]
    emp_list.append(d)
file.close()
print(emp_list)
'''
for emp in emp_list:
    if emp['designation'] == "Developer":
        emp['designation'] = "manager"
    elif emp['designation'] == "manager":
        emp['designation'] = "product owner"
    elif emp['designation'] == "Tester":
        emp['designation'] = "Developer"

with open("emp_data.txt", 'w') as f:
    for emp in emp_list:
        f.write(" ".join(emp.values()))
        f.write(' \n')
'''
book = openpyxl.Workbook()
sheet = book.active
#sheet['A1'] = "Name"
#sheet['B1'] = "Age"
for i in range(1,(len(key)+1)):
    sheet.cell(row=1, column=i).value = key[i-1]
for i in range(2, (len(emp_list)+2)):
    for j in range(1, (len(key) + 1)):
        sheet.cell(row=i, column=j).value = emp_list[i-2][key[j-1]]

book.save("emp_details1.xlsx")

book = openpyxl.load_workbook('emp_details.xlsx')
sheet = book.active
print(sheet['A1'].value)
print(sheet['C2'].value)
