from openpyxl import Workbook

workbook = Workbook()
worksheet = workbook.active
worksheet1 = workbook.create_sheet("worksheet1")
worksheet2 = workbook.create_sheet("worksheet2")
data1 = [['ABC',50,'9876543210'],
      ['PQR',20,'9876454321'],
      ['XYZ',25,'9845454540'],
      ['AMV',35,'9875454510'],
      ['RVT',40,'9876545444']]

data2 = [{'name':'ABC','age':50,'PNo.':'9876543210'},
         {'name':'PQR','age':20,'PNo.':'9876543210'},
         {'name':'AMV','age':25,'PNo.':'9876543210'},
         {'name':'XYZ','age':35,'PNo.':'9876543210'},
         {'name':'RVT','age':40,'PNo.':'9876543210'}]

keys = ['name', 'age', 'PNo.']
for i in range(len(keys)):
    worksheet1.cell(row=1, column=i+1).value = keys[i]
    worksheet2.cell(row=1, column=i+1).value = keys[i]

for row in range(2,len(data1)+2):
    for col in range(1,len(keys)+1):
        worksheet1.cell(row=row, column=col).value = data1[row-2][col-1]

for row in range(2,7):
    for col in range(1,4):
        worksheet2.cell(row=row, column=col).value = data2[row-2][keys[col-1]]

workbook.save("sample_xl2.xlsx")