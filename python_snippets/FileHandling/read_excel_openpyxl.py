import openpyxl

book = openpyxl.load_workbook('emp_data_p6_final.xlsx')
sheet = book.active
emp_list = list()
for row in range(2, sheet.max_row+1):
    emp_data = dict()
    for col in range(1, sheet.max_column+1):
        emp_data[header_list[col-1]]=(sheet.cell(row=row, column=col).value)

    emp_list.append(emp_data)
print(emp_list)