import openpyxl
import pymysql
import time



'''
**********************************   To create excel file and write data in to it

book = openpyxl.Workbook()
sheet = book.active
sheet['A1'] = 56
sheet['A2'] = 43
now = time.strftime("%x")
sheet['A3'] = now
book.save("sample.xlsx")
'''

book = openpyxl.load_workbook('sample.xls')

sheet = book.active
'''
To Access Cell Data

a1 = sheet['A1']
a2 = sheet['A2']
a3 = sheet.cell(row=3, column=1)

cells = sheet['A1': 'B6']

for c1, c2 in cells:
    #print(c1.value, c2.value)

print(sheet.cell(row=1, column=1).value) # to read Value
sheet.cell(row=1, column=1).value = 5 # to write Value

'''


rowTuple = sheet['A2': 'D6']

outfile = open('sample.txt','w')

db = pymysql.connect("localhost", "eknath", "1nath@Mysql", "test")
cursor = db.cursor()

for r in rowTuple:
    if r[3].value < r[2].value:
        print("dump in database")
        outfile.writelines("name : "+r[0].value+" , Remaining  fees of Rs. : "+str(r[2].value - r[3].value)+"\n")
    else:
        print("dump in text")
        print("insert into student_info (name,course,confirmation) values ('"+r[0].value+"','"+r[1].value+"','yes')")
        cursor.execute("insert into student_info (name,course,confirmation) values ('"+r[0].value+"','"+r[1].value+"','yes')")

        print("name : {} , Remained with fees of Rs. : {}".format(r[0].value,r[3].value))


db.commit()
