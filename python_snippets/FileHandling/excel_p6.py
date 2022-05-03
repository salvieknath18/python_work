import openpyxl

emp_data = list()
with open('sampl.txt') as f:
    for line in f.readlines():
        emp_data.append(
            {
                'name': line.strip().split()[0],
                'company': line.strip().split()[1],
                'salary': line.strip().split()[2],
                'designation': line.strip().split()[3],
            }
        )
        line.strip().split()

print(emp_data)

book = openpyxl.Workbook()
sheet = book.active

column_index = 0
for key, value in emp_data[0].items():
    sheet.cell(row=1, column=column_index+1).value = key
    column_index += 1

for row_index, emp in enumerate(emp_data):
    column_index = 0
    for key, value in emp.items():
        sheet.cell(row=row_index+2, column=column_index+1).value = value
        column_index +=1

book.save("emp_data_p6_final.xlsx")
