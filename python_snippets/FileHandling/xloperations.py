import openpyxl
# https://openpyxl.readthedocs.io/en/stable/tutorial.html
import pymysql
import time



'''
**********************************   To create excel file and write data in to it
'''
book = openpyxl.Workbook()
sheet = book.active
sheet['A1'] = 56
sheet['A2'] = 43
now = time.strftime("%x")
sheet['A3'] = now
book.save("sample_p6.xlsx")


book = openpyxl.load_workbook('sample2.xlsx')

sheet = book.active
'''
To Access Cell Data
'''
a1 = sheet['A1']
a2 = sheet['A2']
a3 = sheet.cell(row=3, column=1)

cells = sheet['A1': 'B6']

for c1, c2 in cells:
    print(c1.value, c2.value)

print(sheet.cell(row=1, column=1).value) # to read Value
sheet.cell(row=1, column=1).value = 5 # to write Value

'''
rowTuple = sheet['A2': 'D6']

outfile = open('sample.txt','w')

for r in rowTuple:
    outfile.writelines("name : "+r[0].value+" , Remaining  fees of Rs. : "+str(r[2].value - r[3].value)+"\n")
'''