from openpyxl import Workbook, load_workbook
l1 = list()
workbook = load_workbook("sample_xl.xlsx")
worksheet = workbook.active
#worksheet2['B2'] = "ABCD-1245"
for row in range(2,5):
    for col in range(1,6):
        print(row-2, col-1, end=": ")
        #l1[row-2][col-1] = worksheet.cell(row=row, column=col).value
        print(worksheet.cell(row=row, column=col).value)